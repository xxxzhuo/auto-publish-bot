#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""All-in-one script for 自动发布机器人.

This file merges the previous modules into a single runnable Python file:
- config
- parser / offer_parser / detailed_parser / hkstock_parser / universal_parser
- excel_generator
- login / uploader
- backend API (FastAPI)
- main pipelines
"""

import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import requests
import uvicorn
from fastapi import FastAPI, Header, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl import Workbook
from pydantic import BaseModel


# ============================================================================
# Config
# ============================================================================

# 🦞 先搜精选 v4.1 - 删除默认手机号，必须动态传入
# 不再设置默认值，强制要求用户提供手机号
REQUIRED_PHONE_NUMBER = None  # 必须动态传入
REQUIRED_SMS_CODE = "2222"  # 默认验证码，可被覆盖

# 兼容旧代码（但不再使用默认值）
PHONE_NUMBER = None
SMS_CODE = REQUIRED_SMS_CODE

BASE_URL = "http://129.204.124.204:89/dev-api"
LOGIN_API = "/api/mts-api/login"
UPLOAD_API = "/api/mts-api/clue/import"

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = str(BASE_DIR)
PRODUCTS_DIR = str(BASE_DIR / "products")
OUTPUT_DIR = str(BASE_DIR / "output")
LOGS_DIR = str(BASE_DIR / "logs")

TIMEOUT = 30
RETRY_COUNT = 3

# SIMULATE: only parse + excel
# REAL: login + upload
RUN_MODE = "REAL"


# ============================================================================
# 品牌名映射（英文 -> 中文）🦞 先搜精选 - 强制映射
# ============================================================================

BRAND_NAME_MAPPING = {
    # 英文品牌名 -> 中文品牌名
    "Samsung": "三星",
    "SAMSUNG": "三星",
    "samsung": "三星",
    "SK Hynix": "海力士",
    "SKHynix": "海力士",
    "Hynix": "海力士",
    "hynix": "海力士",
    "Micron": "镁光",
    "micron": "镁光",
    "Nanya": "南亚",
    "nanya": "南亚",
    "Nanya Technology": "南亚科技",
    "Kingston": "金士顿",
    "kingston": "金士顿",
    "GigaDevice": "兆易创新",
    "gigadevice": "兆易创新",
    "Winbond": "华邦",
    "winbond": "华邦",
    "Puya": "晶存",
    "puya": "晶存",
    "Intel": "英特尔",
    "intel": "英特尔",
    "INTEL": "英特尔",
}

# 已知品牌前缀映射（用于强制映射）
BRAND_PREFIX_MAPPING = {
    "K4": "三星",
    "K3": "三星",
    "M32": "三星",
    "KLM": "三星",
    "H5": "海力士",
    "H9": "海力士",
    "MT4": "镁光",
    "MT5": "镁光",
    "MT6": "镁光",
    "MT29": "镁光",
    "NT5": "南亚",
    "NT6": "南亚",
    "D": "金士顿",
    "KVR": "金士顿",
    "GD": "兆易创新",
    "W25": "华邦",
    "W94": "华邦",
    "P25": "晶存",
    "29F": "英特尔",
}


def map_brand_name(brand: str, part_number: str = None) -> str:
    """强制将品牌名映射为中文
    
    Args:
        brand: 品牌名（可能是英文或中文）
        part_number: 料号型号（用于辅助判断）
        
    Returns:
        映射后的中文品牌名，强制返回中文
    """
    if not brand:
        # 如果品牌为空，尝试从料号推断
        if part_number:
            for prefix, chinese_name in BRAND_PREFIX_MAPPING.items():
                if part_number.startswith(prefix):
                    return chinese_name
        return "未知品牌"
    
    brand = brand.strip()
    
    # 🦞 先搜精选 - 强制映射为中文
    # 1. 先尝试精确匹配
    if brand in BRAND_NAME_MAPPING:
        return BRAND_NAME_MAPPING[brand]
    
    # 2. 尝试忽略大小写匹配
    brand_upper = brand.upper()
    for key, value in BRAND_NAME_MAPPING.items():
        if key.upper() == brand_upper:
            return value
    
    # 3. 如果已经是中文，直接返回
    if any('\u4e00' <= c <= '\u9fff' for c in brand):
        return brand
    
    # 4. 未知英文品牌，返回原值（但这种情况应该很少）
    return brand


# ============================================================================
# Common constants
# ============================================================================

EXCEL_COLUMNS = [
    "企业名称",
    "货主",
    "性别",
    "手机号",
    "微信",
    "产品名称",
    "品牌",
    "料号型号",
    "单价",
    "货币单位（usd/cny）",
    "数量",
    "单位",
    "货所在地",
    "批次（DC）",
    "货物情况（多选）",
    "产品分类",
    "应用",
    "起订量",
    "有效期",
    "交货天数",
    "厂家",
    "重量",
    "尺寸 - 长",
    "尺寸 - 宽",
    "尺寸 - 高",
    "替代品牌",
    "替代型号",
]


def build_empty_product() -> Dict:
    """Create a standard product dict with all target fields."""
    return {
        "企业名称": None,
        "货主": None,
        "性别": None,
        "手机号": None,
        "微信": None,
        "产品名称": None,
        "品牌": None,
        "料号型号": None,
        "单价": None,
        "货币单位（usd/cny）": "usd",
        "数量": None,
        "单位": "个",
        "货所在地": "HK",
        "批次（DC）": None,
        "货物情况（多选）": "全新",
        "产品分类": "存储芯片",
        "应用": None,
        "起订量": None,
        "有效期": None,
        "交货天数": 7,
        "厂家": None,
        "重量": None,
        "尺寸 - 长": None,
        "尺寸 - 宽": None,
        "尺寸 - 高": None,
        "替代品牌": None,
        "替代型号": None,
    }


# ============================================================================
# Parser: txt key-value files
# ============================================================================


class ProductParser:
    """产品信息解析器（按 txt 字段解析）"""

    FIELD_MAPPING = {
        "产品名称": "产品名称",
        "产品名": "产品名称",
        "名称": "产品名称",
        "品牌": "品牌",
        "品牌名": "品牌",
        "型号": "料号型号",
        "料号": "料号型号",
        "料号型号": "料号型号",
        "part": "料号型号",
        "part_number": "料号型号",
        "单价": "单价",
        "价格": "单价",
        "price": "单价",
        "货币": "货币单位（usd/cny）",
        "货币单位": "货币单位（usd/cny）",
        "currency": "货币单位（usd/cny）",
        "数量": "数量",
        "qty": "数量",
        "quantity": "数量",
        "单位": "单位",
        "货所在地": "货所在地",
        "所在地": "货所在地",
        "地点": "货所在地",
        "批次": "批次（DC）",
        "dc": "批次（DC）",
        "date_code": "批次（DC）",
        "货物情况": "货物情况（多选）",
        "成色": "货物情况（多选）",
        "情况": "货物情况（多选）",
        "产品分类": "产品分类",
        "分类": "产品分类",
        "应用": "应用",
        "起订量": "起订量",
        "moq": "起订量",
        "有效期": "有效期",
        "交货天数": "交货天数",
        "交期": "交货天数",
        "厂家": "厂家",
        "厂商": "厂家",
        "重量": "重量",
        "尺寸": "尺寸 - 长",
        "长": "尺寸 - 长",
        "宽": "尺寸 - 宽",
        "高": "尺寸 - 高",
        "替代品牌": "替代品牌",
        "替代": "替代品牌",
        "替代型号": "替代型号",
        "企业名称": "企业名称",
        "公司": "企业名称",
        "货主": "货主",
        "姓名": "货主",
        "性别": "性别",
        "手机号": "手机号",
        "电话": "手机号",
        "微信": "微信",
        "wechat": "微信",
    }

    def parse_file(self, file_path: str) -> Optional[Dict]:
        if not os.path.exists(file_path):
            print(f"❌ 文件不存在：{file_path}")
            return None

        product_data = {col: None for col in EXCEL_COLUMNS}

        try:
            with open(file_path, "r", encoding="utf-8") as f:
                content = f.read()

            for line in content.strip().split("\n"):
                line = line.strip()
                if not line or line.startswith("#"):
                    continue

                fullwidth_colon = chr(0xFF1A)
                halfwidth_colon = chr(0x003A)

                if fullwidth_colon in line:
                    key, value = line.split(fullwidth_colon, 1)
                elif halfwidth_colon in line:
                    key, value = line.split(halfwidth_colon, 1)
                elif "=" in line:
                    key, value = line.split("=", 1)
                else:
                    continue

                key = key.strip()
                value = value.strip()

                if not key or not value:
                    continue

                excel_field = self.FIELD_MAPPING.get(key)
                if excel_field:
                    product_data[excel_field] = value
                else:
                    matched = self.fuzzy_match(key)
                    if matched:
                        product_data[matched] = value

            if product_data["货币单位（usd/cny）"] is None:
                product_data["货币单位（usd/cny）"] = "cny"
            if product_data["单位"] is None:
                product_data["单位"] = "个"

            return product_data

        except Exception as e:
            print(f"❌ 解析失败 {file_path}: {e}")
            return None

    def fuzzy_match(self, key: str) -> Optional[str]:
        key_lower = key.lower()
        for txt_key, excel_key in self.FIELD_MAPPING.items():
            if txt_key.lower() in key_lower or key_lower in txt_key.lower():
                return excel_key
        return None

    def parse_directory(self, directory: Optional[str] = None) -> List[Dict]:
        directory = directory or PRODUCTS_DIR

        if not os.path.exists(directory):
            print(f"❌ 目录不存在：{directory}")
            return []

        txt_files = list(Path(directory).glob("*.txt"))
        if not txt_files:
            print(f"⚠️  目录中没有找到 txt 文件：{directory}")
            return []

        products: List[Dict] = []
        for txt_file in txt_files:
            product_data = self.parse_file(str(txt_file))
            if product_data:
                products.append(product_data)
        return products


# ============================================================================
# Excel generator
# ============================================================================


class ExcelGenerator:
    """Excel 表格生成器"""

    COLUMNS = EXCEL_COLUMNS

    def __init__(self):
        self.wb = None
        self.ws = None

    def create_excel(self, products: List[Dict], filename: Optional[str] = None) -> Optional[str]:
        if not products:
            print("❌ 没有产品数据")
            return None

        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "供需信息"

        self._write_headers()
        self._write_data(products)
        self._auto_adjust_columns()

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"供需信息_{timestamp}.xlsx"

        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)

        filepath = os.path.join(OUTPUT_DIR, filename)
        self.wb.save(filepath)

        return filepath

    def _write_headers(self):
        for col, header in enumerate(self.COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    def _write_data(self, products: List[Dict]):
        for row_idx, product in enumerate(products, 2):
            for col_idx, field in enumerate(self.COLUMNS, 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.value = product.get(field)
                cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center")

    def _auto_adjust_columns(self):
        column_widths = {
            "企业名称": 20,
            "货主": 10,
            "性别": 6,
            "手机号": 13,
            "微信": 15,
            "产品名称": 25,
            "品牌": 15,
            "料号型号": 20,
            "单价": 10,
            "货币单位（usd/cny）": 12,
            "数量": 8,
            "单位": 6,
            "货所在地": 10,  # 大陆/香港，缩短宽度
            "批次（DC）": 12,
            "货物情况（多选）": 15,
            "产品分类": 15,
            "应用": 15,
            "起订量": 8,
            "有效期": 12,
            "交货天数": 8,
            "厂家": 15,
            "重量": 8,
            "尺寸 - 长": 8,
            "尺寸 - 宽": 8,
            "尺寸 - 高": 8,
            "替代品牌": 15,
            "替代型号": 15,
        }

        for col_idx, header in enumerate(self.COLUMNS, 1):
            width = column_widths.get(header, 12)
            letter = openpyxl.utils.get_column_letter(col_idx)
            self.ws.column_dimensions[letter].width = width

    def append_to_existing(self, filepath: str, products: List[Dict]) -> bool:
        if not os.path.exists(filepath):
            print(f"❌ 文件不存在：{filepath}")
            return False

        self.wb = openpyxl.load_workbook(filepath)
        if "供需信息" not in self.wb.sheetnames:
            print("❌ 文件中没有'供需信息'工作表")
            return False

        self.ws = self.wb["供需信息"]
        start_row = self.ws.max_row + 1

        for row_idx, product in enumerate(products, start_row):
            for col_idx, field in enumerate(self.COLUMNS, 1):
                self.ws.cell(row=row_idx, column=col_idx).value = product.get(field)

        self.wb.save(filepath)
        return True


# ============================================================================
# Login & uploader
# ============================================================================


class LoginManager:
    """登录管理器（真实接口）"""

    def __init__(self):
        self.session = requests.Session()
        self.token = None
        self.user_info = None

    def login(self, mobile: Optional[str] = None, sms_code: Optional[str] = None) -> bool:
        # 🦞 先搜精选 v4.1 - 强制要求动态手机号，不再使用默认值
        if not mobile:
            print("❌ 错误：必须提供手机号")
            return False
        
        sms_code = sms_code or REQUIRED_SMS_CODE

        if not mobile or not sms_code:
            print("❌ 错误：请配置手机号和短信验证码")
            return False

        url = f"{BASE_URL}{LOGIN_API}"
        payload = {"mobile": mobile, "smsCode": sms_code}

        for attempt in range(1, RETRY_COUNT + 1):
            try:
                response = self.session.post(url, json=payload, timeout=TIMEOUT)

                if response.status_code == 200:
                    result = response.json()
                    if result.get("code") == 200:
                        data = result.get("data", {})
                        self.token = data.get("token")
                        self.user_info = {
                            "username": data.get("username"),
                            "nickname": data.get("nickname"),
                            "mobile": data.get("mobile"),
                            "userId": data.get("userId"),
                        }
                        return True

                    return False

            except requests.exceptions.RequestException as e:
                if attempt == RETRY_COUNT:
                    print(f"❌ 登录请求失败：{e}")

        return False

    def get_headers(self) -> Dict:
        headers = {
            "Content-Type": "application/json",
            "User-Agent": "AutoBot/2.0",
        }
        if self.token:
            headers["Authorization"] = f"Bearer {self.token}"
        return headers

    def get_session(self) -> requests.Session:
        return self.session

    def get_token(self) -> Optional[str]:
        return self.token


class Uploader:
    """文件上传器（真实接口）"""

    def __init__(self, session: Optional[requests.Session] = None, headers: Optional[Dict] = None):
        self.session = session or requests.Session()
        self.headers = headers or {}
        self.last_error: Optional[str] = None

    def upload_excel(self, filepath: str) -> bool:
        if not os.path.exists(filepath):
            self.last_error = f"文件不存在: {filepath}"
            return False

        url = f"{BASE_URL}{UPLOAD_API}"
        self.last_error = None

        for attempt in range(1, RETRY_COUNT + 1):
            try:
                with open(filepath, "rb") as f:
                    files = {
                        "file": (
                            os.path.basename(filepath),
                            f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    }
                    upload_headers = {k: v for k, v in self.headers.items() if k != "Content-Type"}
                    response = self.session.post(
                        url,
                        files=files,
                        headers=upload_headers,
                        timeout=TIMEOUT * 2,
                    )

                if response.status_code == 200:
                    result = response.json()
                    if result.get("code") == 200:
                        return True
                    self.last_error = result.get("msg", "未知错误")
                    return False

                self.last_error = f"HTTP {response.status_code}: {response.text[:200]}"

            except requests.exceptions.RequestException as e:
                self.last_error = str(e)

        return False


# ============================================================================
# OFFER parsers
# ============================================================================


class OfferParser:
    """鸿达格式 OFFER 文本解析器"""

    DEFAULT_BRAND_MAP = {
        # 南亚科技 (Nanya)
        "NT5A": "南亚科技 (Nanya)",
        "NT5B": "南亚科技 (Nanya)",
        "NT5C": "南亚科技 (Nanya)",
        "NT5D": "南亚科技 (Nanya)",
        "NT5L": "南亚科技 (Nanya)",
        "NT5U": "南亚科技 (Nanya)",
        "NT6A": "南亚科技 (Nanya)",
        "NT6B": "南亚科技 (Nanya)",
        "NT6C": "南亚科技 (Nanya)",
        "NT6U": "南亚科技 (Nanya)",
        # SK Hynix
        "H5C": "SK Hynix",
        "H5AN": "SK Hynix",
        "H5HC": "SK Hynix",
        "H5T4": "SK Hynix",
        "H5TQ": "SK Hynix",
        "H9C": "SK Hynix",
        "H9HC": "SK Hynix",
        "H9TQ": "SK Hynix",
        # 镁光 (Micron)
        "MT4": "Micron",
        "MT40": "Micron",
        "MT41": "Micron",
        "MT41K": "Micron",
        "MT41J": "Micron",
        "MT42": "Micron",
        "MT46": "Micron",
        "MT5": "Micron",
        "MT53": "Micron",
        "MT53E": "Micron",
        "MT58": "Micron",
        "MT6": "Micron",
        "MT62": "Micron",
        "MT62F": "Micron",
        "MT29": "Micron",
        "MT29F": "Micron",
        "25Q": "Micron",
        "29F": "Micron",
        # 三星 (Samsung)
        "M32": "Samsung",
        "K4": "Samsung",
        "K4A": "Samsung",
        "K4B": "Samsung",
        "K4U": "Samsung",
        "K4F": "Samsung",
        "K4G": "Samsung",
        "K4L": "Samsung",
        "K8": "Samsung",
        "K8E": "Samsung",
        "K8G": "Samsung",
        "KLM": "Samsung",
        "KMF": "Samsung",
        "KLUD": "Samsung",
        "HN8T": "Samsung",
        "H26T": "Samsung",
        "THG": "Samsung",
        "K4AAG": "Samsung",
        "K3K": "Samsung",
        "K3U": "Samsung",
        "KMZ": "Samsung",
        # 金士顿 (Kingston)
        "D": "Kingston",
        "D1": "Kingston",
        "D2": "Kingston",
        "D3": "Kingston",
        "D4": "Kingston",
        "KVR": "Kingston",
        # 兆易创新 (GigaDevice)
        "GD": "GigaDevice",
        "GD25": "GigaDevice",
        "GD25L": "GigaDevice",
        "GD25Q": "GigaDevice",
        "GD25W": "GigaDevice",
        # 华邦 (Winbond)
        "W": "Winbond",
        "W25": "Winbond",
        "W25Q": "Winbond",
        "W25N": "Winbond",
        "W94": "Winbond",
        "W98": "Winbond",
        # 晶存 (Puya)
        "P25": "晶存 (Puya)",
        "P25D": "晶存 (Puya)",
        "P25Q": "晶存 (Puya)",
        # 英特尔 (Intel)
        "29F": "Intel",
        "PC29": "Intel",
        "SSD": "Intel",
        # 美光 (Micron 别名)
        "MICRON": "Micron",
        "MT": "Micron",
    }

    COMPANY_PATTERNS = [
        r"^(.+?)\s*OFFER[:：]?\s*$",
        r"^(.+?)\s*现货[:：]?\s*$",
        r"^(.+?)\s*实单[:：]?\s*$",
        r"^(.+?)\s*开价[:：]?\s*$",
    ]

    def parse_offer_text(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        lines = text.strip().split("\n")
        company = self._extract_company(lines)

        for line in lines:
            line = line.strip()
            if not line or self._is_title_line(line):
                continue

            product = self._parse_line(line, company)
            if product:
                products.append(product)

        return products

    def _extract_company(self, lines: List[str]) -> Optional[str]:
        for line in lines[:5]:
            line = line.strip()
            for pattern in self.COMPANY_PATTERNS:
                match = re.search(pattern, line, re.IGNORECASE)
                if match:
                    company = match.group(1).strip()
                    company = re.sub(r"^[出发][，,!]", "", company)
                    return company
        return None

    def _is_title_line(self, line: str) -> bool:
        title_keywords = ["offer", "现货", "实单", "开价", "价格", "出，", "发，"]
        line_lower = line.lower()
        return any(keyword in line_lower for keyword in title_keywords)

    def _parse_line(self, line: str, company: Optional[str] = None) -> Optional[Dict]:
        product = build_empty_product()
        product["企业名称"] = company
        product["货所在地"] = "HK"
        product["交货天数"] = 7

        part_patterns = [
            r"(NT\d+[A-Z]+\d+[A-Z]*[-:]\w+)",
            r"(H5C\d+[A-Z]+\d+[A-Z])",
            r"(H5AN\d+[A-Z]+[-:]\w+)",
            r"(MT\d+[A-Z]+\d+[A-Z]*[-:]\d+[A-Z]:\w+)",
            r"(MT\d+[A-Z]+\d+[A-Z]*[-:]\d+[A-Z]\s*\w+:\w+)",
            r"(M32\d+[A-Z]+\d+[A-Z]*[-:]\w+)",
        ]

        part_number = None
        for pattern in part_patterns:
            match = re.search(pattern, line)
            if match:
                part_number = match.group(1).replace(" ", "")
                break

        if not part_number:
            match = re.match(r"^([A-Z]+\d+[A-Z0-9\-:]+)", line.strip())
            if match:
                part_number = match.group(1)

        if not part_number:
            return None

        product["料号型号"] = part_number
        product["产品名称"] = part_number

        brand = self._identify_brand(part_number)
        # 🦞 先搜精选 v4.2 - 强制映射为中文（带 part_number 参数）
        product["品牌"] = map_brand_name(brand, part_number)
        product["厂家"] = map_brand_name(brand, part_number)

        price_match = re.search(r"usd\s*(\d+\.?\d*)", line, re.IGNORECASE)
        if price_match:
            product["单价"] = float(price_match.group(1))

        dc_match = re.search(r"(?:dc\s*)?(\d{2}\+)", line, re.IGNORECASE)
        if dc_match:
            product["批次（DC）"] = dc_match.group(1)

        qty_match = re.search(r"\*\s*(\d+)\s*pcs", line, re.IGNORECASE)
        if qty_match:
            product["数量"] = int(qty_match.group(1))

        line_lower = line.lower()
        if "1week" in line_lower or "1 week" in line_lower:
            product["交货天数"] = 7
        elif "2days" in line_lower or "2 days" in line_lower:
            product["交货天数"] = 2
        elif "next monday" in line_lower:
            product["交货天数"] = 14
        elif "stk" in line_lower:
            product["交货天数"] = 1

        if "non cn" in line_lower:
            product["货所在地"] = "HK (Non-CN)"
        elif "hk" in line_lower:
            product["货所在地"] = "HK"
        elif "partial" in line_lower:
            product["货物情况（多选）"] = "部分"

        if "reel" in line_lower:
            product["货物情况（多选）"] = "全新 (Reel)"
        if "around" in line_lower:
            product["货物情况（多选）"] = "现货 around"

        return product

    def _identify_brand(self, part_number: str) -> str:
        for prefix, brand in self.DEFAULT_BRAND_MAP.items():
            if part_number.startswith(prefix):
                return brand
        return "未知品牌"

    def parse_offer_file(self, filepath: str) -> List[Dict]:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                return self.parse_offer_text(f.read())
        except Exception:
            return []


class DetailedLineParser:
    """单行详细格式解析器"""

    BRAND_PATTERNS = {
        # 南亚科技 (Nanya)
        "南亚科技 (Nanya)": [
            r"^NT5A", r"^NT5B", r"^NT5C", r"^NT5D", r"^NT5L", r"^NT5U",
            r"^NT6A", r"^NT6B", r"^NT6C", r"^NT6U",
            r"南亚", r"Nanya",
        ],
        # SK Hynix
        "SK Hynix": [
            r"^H5C", r"^H5AN", r"^H5HC", r"^H5T4", r"^H5TQ",
            r"^H9C", r"^H9HC", r"^H9TQ",
            r"海力士", r" SK ", r"Hynix",
        ],
        # 镁光 (Micron)
        "Micron": [
            r"^MT4\b", r"^MT40", r"^MT41", r"^MT41K", r"^MT41J", r"^MT42", r"^MT46",
            r"^MT5\b", r"^MT53", r"^MT53E", r"^MT58",
            r"^MT6\b", r"^MT62", r"^MT62F",
            r"^MT29", r"^MT29F",
            r"^25Q", r"^29F",
            r"镁光", r"美光", r"Micron",
        ],
        # 三星 (Samsung)
        "Samsung": [
            r"^M32",
            r"^K4\b", r"^K4A", r"^K4B", r"^K4U", r"^K4F", r"^K4G", r"^K4L",
            r"^K8\b", r"^K8E", r"^K8G",
            r"^KLM", r"^KMF", r"^KLUD",
            r"^HN8T", r"^H26T", r"^THG",
            r"^K4AAG", r"^K3K", r"^K3U", r"^KMZ",
            r"三星", r"Samsung",
        ],
        # 金士顿 (Kingston)
        "Kingston": [
            r"^D\d", r"^D1", r"^D2", r"^D3", r"^D4",
            r"^KVR",
            r"金士顿", r"金士", r"Kingston",
        ],
        # 兆易创新 (GigaDevice)
        "GigaDevice": [
            r"^GD\b", r"^GD25", r"^GD25L", r"^GD25Q", r"^GD25W",
            r"兆易", r"GigaDevice",
        ],
        # 华邦 (Winbond)
        "Winbond": [
            r"^W\b", r"^W25", r"^W25Q", r"^W25N", r"^W94", r"^W98",
            r"华邦", r"Winbond",
        ],
        # 晶存 (Puya)
        "晶存 (Puya)": [
            r"^P25\b", r"^P25D", r"^P25Q",
            r"晶存", r"Puya",
        ],
        # 英特尔 (Intel)
        "Intel": [
            r"^29F", r"^PC29", r"^SSD",
            r"英特尔", r"Intel",
        ],
    }

    PRODUCT_TYPES = {
        "LPDDR4X": "LPDDR4X",
        "LPDDR4": "LPDDR4",
        "LPDDR5": "LPDDR5",
        "LPDDR3": "LPDDR3",
        "DDR4": "DDR4",
        "DDR3": "DDR3",
        "EMMC": "eMMC",
        "UFS": "UFS",
        "NAND": "NAND",
    }

    def parse(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        lines = text.strip().split("\n")

        start_idx = 0
        for i, line in enumerate(lines):
            lower = line.lower()
            if "offer" in lower or "现货" in lower or "下单" in lower:
                start_idx = i + 1
                break

        for i in range(start_idx, len(lines)):
            line = lines[i].strip()
            if not line:
                continue
            lower = line.lower()
            if "offer" in lower or "现货" in lower:
                continue

            product = self._parse_line(line)
            if product and product.get("料号型号"):
                products.append(product)

        return products

    def _parse_line(self, line: str) -> Optional[Dict]:
        product = build_empty_product()
        product["企业名称"] = "鸿达半导体"

        part_match = re.match(r"^([A-Z0-9\-]+)", line)
        if part_match:
            product["料号型号"] = part_match.group(1)
            product["产品名称"] = part_match.group(1)

        if not product["料号型号"]:
            return None

        brand = self._identify_brand(line)
        # 🦞 先搜精选 - 品牌名映射（中文->英文）
        product["品牌"] = map_brand_name(brand, part_number)
        product["厂家"] = map_brand_name(brand, part_number)

        upper_line = line.upper()
        for type_name, type_value in self.PRODUCT_TYPES.items():
            if type_name in upper_line:
                product["产品分类"] = type_value
                break

        capacity_match = re.search(r"(\d+)\s*(GB|MB|TB)", line, re.IGNORECASE)
        if capacity_match:
            product["数量"] = int(capacity_match.group(1))
            product["单位"] = capacity_match.group(2).upper()

            remaining = line[capacity_match.end() :]
            dc_match = re.search(r"(\d{2})\s*\+", remaining)
            if dc_match:
                product["批次（DC）"] = dc_match.group(1) + "+"

        price_match = re.search(r"(\d+\.?\d*)\s*u\b", line, re.IGNORECASE)
        if price_match:
            try:
                product["单价"] = float(price_match.group(1))
            except Exception:
                pass

        if "现货" in line:
            product["货所在地"] = "HK"
            product["交货天数"] = 1
            product["货物情况（多选）"] = "现货"

        return product

    def _identify_brand(self, line: str) -> str:
        for brand, patterns in self.BRAND_PATTERNS.items():
            for pattern in patterns:
                if re.search(pattern, line, re.IGNORECASE):
                    return brand

        part_match = re.match(r"^([A-Z]{2,}\d+)", line)
        if part_match:
            prefix = part_match.group(1)[:4]
            if prefix.startswith("K4") or prefix.startswith("KLM"):
                return "Samsung"
            if prefix.startswith("H9C") or prefix.startswith("H5C"):
                return "SK Hynix"
            if prefix.startswith("MT"):
                return "Micron"
            if prefix.startswith("NT"):
                return "南亚科技 (Nanya)"

        return "未知品牌"


class HKOfferParser:
    """香港 OFFER 格式解析器 (2026-01-08 HK OFFER 格式)"""

    BRAND_MAP = {
        "SAMSUNG": "Samsung",
        "SK HYNIX": "SK Hynix",
        "SK Hynix": "SK Hynix",
        "NANYA": "南亚科技 (Nanya)",
        "Nanya": "南亚科技 (Nanya)",
        "MICRON": "Micron",
        "H5AN": "SK Hynix",
        "H5TQ": "SK Hynix",
        "K4": "Samsung",
        "K8": "Samsung",
        "KLM": "Samsung",
        "H26": "SK Hynix",
        "MT40": "Micron",
        "MT41": "Micron",
        "NT5": "南亚科技 (Nanya)",
        "NT6": "南亚科技 (Nanya)",
        "H9HC": "SK Hynix",
        "K4F": "Samsung",
        "K4U": "Samsung",
        "CXDB": "CXMT",
        "SDIN": "SanDisk",
    }

    def parse(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        lines = text.strip().split("\n")
        
        current_brand = None
        current_category = None
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 检测品牌分类行 (如 SAMSUNG DDR3, SK Hynix DDR4)
            category_match = re.match(r'^(SAMSUNG|SK HYNIX|SK Hynix|NANYA|Nanya|MICRON|CXMT|Sndisk|SanDisk)\s+(DDR[345]|LPDDR[45X]?|EMMC|NAND|SSD)?', line, re.IGNORECASE)
            if category_match:
                current_brand = category_match.group(1).upper()
                current_category = category_match.group(2).upper() if category_match.group(2) else None
                continue
            
            # 跳过标题行
            if 'OFFER' in line.upper() or 'DATE' in line.upper():
                continue
            
            # 解析产品行：规格 料号 价格 [DC]
            # 格式 1: 256*16 K4B4G1646E-BCNB $BID
            # 格式 2: 1024*16 K4AAG165WA-BCWE0CV $80 DC24+
            # 格式 3:  H5AG46DXNDX117N $95 (无规格)
            # 格式 4: MT40A256M16GE-083E:B $4.8 (料号带冒号)
            # 格式 5: MLC V5.1 4GB KLM4G1FETE-B041001 $11.2 (eMMC 格式)
            # 格式 6: 1GB K4U8E3S4AD-MGCL0JP $23 (LPDDR4X 容量格式)
            # 格式 7: 32Gb H9HCNNNCPUMLXR-NEE $60 (LPDDR4X Gb 格式)
            # 格式 8: 16Gb CXDB4CBAM-ML-A $24 (CXMT 格式)
            
            part_number = None
            price_str = None
            dc = None
            spec_width = None
            spec_depth = None
            
            # 尝试 eMMC 格式：MLC V5.1 4GB KLM4G1FETE-B041001 $11.2
            emmc_match = re.match(
                r'^(?:MLC\s+)?(?:V\d+\.\d+\s+)?(?:\d+GB\s+)?([A-Z0-9\-]+)\s+\$?([\d.]+)\s*(?:DC(\d+\+))?',
                line,
                re.IGNORECASE
            )
            if emmc_match:
                part_number = emmc_match.group(1)
                price_str = emmc_match.group(2)
                dc = emmc_match.group(3)
            
            # 尝试 LPDDR4X/CXMT 格式：1GB K4U8E3S4AD-MGCL0JP $23 或 32Gb H9HCNNNCPUMLXR-NEE $60
            if not part_number:
                cap_match = re.match(
                    r'^(\d+(?:\.\d+)?(?:GB|Gb))\s+([A-Z0-9\-]+)\s+\$?([\d.]+)\s*(?:DC(\d+\+))?',
                    line,
                    re.IGNORECASE
                )
                if cap_match:
                    spec_width = cap_match.group(1)
                    part_number = cap_match.group(2)
                    price_str = cap_match.group(3)
                    dc = cap_match.group(4)
            
            # 尝试标准格式：256*16 K4B4G1646E-BCNB $BID
            if not part_number:
                std_match = re.match(
                    r'^(?:(\d+)\*(\d+)\s+)?([A-Z0-9\-\:]+)\s+\$?([\d.]+|BID)\s*(?:DC(\d+\+))?',
                    line,
                    re.IGNORECASE
                )
                if std_match:
                    spec_width = std_match.group(1)
                    spec_depth = std_match.group(2)
                    part_number = std_match.group(3)
                    price_str = std_match.group(4)
                    dc = std_match.group(5)
            
            if not part_number or not price_str:
                continue
            
            # 跳过纯数字或太短的料号
            if len(part_number) < 6 or part_number.isdigit():
                continue
            
            # 确定品牌
            brand = self._identify_brand(part_number, current_brand)
            
            # 创建产品
            product = build_empty_product()
            product["企业名称"] = "香港现货供应商"
            product["料号型号"] = part_number
            product["产品名称"] = f"{brand} {current_category or '存储芯片'} {part_number}"
            # 🦞 先搜精选 - 品牌名映射（中文->英文）
            product["品牌"] = map_brand_name(brand, part_number)
            product["厂家"] = map_brand_name(brand, part_number)
            product["产品分类"] = current_category or "存储芯片"
            
            # 价格
            if price_str.upper() == "BID":
                product["单价"] = None  # BID 表示询价
            else:
                try:
                    product["单价"] = float(price_str)
                except:
                    pass
            
            # 批次
            if dc:
                product["批次（DC）"] = dc + "+"
            
            # 货源地映射：英文/缩写 → 大陆/香港
            # 规则：
            # 1. CXMT (长鑫) 是大陆厂商 → 大陆
            # 2. 标题含"HK OFFER"或"香港现货" → 香港
            # 3. 其他默认 → 香港
            if brand == "CXMT":
                product["货所在地"] = "大陆"
            elif "HK OFFER" in text.upper() or "香港现货" in text:
                product["货所在地"] = "香港"
            else:
                product["货所在地"] = "香港"
            
            # 规格转换
            if spec_width and spec_depth:
                product["数量"] = int(spec_width)
                product["单位"] = f"{spec_depth}bit"
            
            products.append(product)
        
        return products
    
    def _identify_brand(self, part_number: str, context_brand: Optional[str] = None) -> str:
        # 先根据料号前缀识别
        for prefix, brand in self.BRAND_MAP.items():
            if part_number.upper().startswith(prefix):
                return brand
        
        # 如果料号无法识别，使用上下文品牌
        if context_brand:
            brand_map = {
                "SAMSUNG": "Samsung",
                "SK HYNIX": "SK Hynix",
                "NANYA": "南亚科技 (Nanya)",
                "MICRON": "Micron",
                "CXMT": "CXMT",
            }
            return brand_map.get(context_brand.upper(), "未知品牌")
        
        return "未知品牌"


class HKStockParser:
    """香港现货格式解析器"""

    BRAND_PATTERNS = {
        # 南亚科技 (Nanya)
        "南亚科技 (Nanya)": [
            r"^NT5A", r"^NT5B", r"^NT5C", r"^NT5D", r"^NT5L", r"^NT5U",
            r"^NT6A", r"^NT6B", r"^NT6C", r"^NT6U",
        ],
        # SK Hynix
        "SK Hynix": [
            r"^H5C", r"^H5AN", r"^H5HC", r"^H5T4", r"^H5TQ",
            r"^H9C", r"^H9HC", r"^H9TQ",
        ],
        # 镁光 (Micron)
        "Micron": [
            r"^MT4\b", r"^MT40", r"^MT41", r"^MT41K", r"^MT41J", r"^MT42", r"^MT46",
            r"^MT5\b", r"^MT53", r"^MT53E", r"^MT58",
            r"^MT6\b", r"^MT62", r"^MT62F",
            r"^MT29", r"^MT29F",
            r"^25Q", r"^29F",
        ],
        # 三星 (Samsung)
        "Samsung": [
            r"^M32",
            r"^K4\b", r"^K4A", r"^K4B", r"^K4U", r"^K4F", r"^K4G", r"^K4L",
            r"^K8\b", r"^K8E", r"^K8G",
            r"^KLM", r"^KMF", r"^KLUD",
            r"^HN8T", r"^H26T", r"^THG",
            r"^K4AAG", r"^K3K", r"^K3U", r"^KMZ",
        ],
        # 金士顿 (Kingston)
        "Kingston": [
            r"^D\d", r"^D1", r"^D2", r"^D3", r"^D4",
            r"^KVR",
        ],
        # 兆易创新 (GigaDevice)
        "GigaDevice": [
            r"^GD\b", r"^GD25", r"^GD25L", r"^GD25Q", r"^GD25W",
        ],
        # 华邦 (Winbond)
        "Winbond": [
            r"^W\b", r"^W25", r"^W25Q", r"^W25N", r"^W94", r"^W98",
        ],
        # 晶存 (Puya)
        "晶存 (Puya)": [
            r"^P25\b", r"^P25D", r"^P25Q",
        ],
        # 英特尔 (Intel)
        "Intel": [
            r"^29F", r"^PC29", r"^SSD",
        ],
    }

    REMARK_KEYWORDS = {
        "小盒开": "小盒装",
        "托盘": "托盘装",
        "涂标": "涂标货",
        "reel": "Reel 装",
        "bulk": "散装",
    }

    def parse(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        lines = self._preprocess_text(text)

        if not lines:
            return products

        company = self._extract_company(lines[0])
        for i, line in enumerate(lines[1:], 1):
            product = self._parse_product_line(line, company)
            if product:
                products.append(product)

        return products

    def _preprocess_text(self, text: str) -> List[str]:
        raw_lines = text.strip().split("\n")
        processed = []
        for line in raw_lines:
            line = line.strip()
            if not line:
                continue
            line = re.sub(r"\s+", " ", line)
            processed.append(line)
        return processed

    def _extract_company(self, first_line: str) -> Optional[str]:
        line = first_line.strip()
        patterns = [
            (r"^(.+?)\s*香港现货", 1),
            (r"^(.+?)\s*实单", 1),
            (r"^(.+?)\s*开价", 1),
            (r"^(.+?)\s*OFFER", 1),
            (r"^[出发][，,!\s]+\s*(.+)", 1),
        ]

        for pattern, group in patterns:
            match = re.search(pattern, line, re.IGNORECASE)
            if match:
                company = match.group(group).strip()
                company = re.sub(r"\s*(香港 | 现货 | 实单 | 开价|OFFER).*$", "", company, flags=re.IGNORECASE)
                company = company.strip()
                return company if company else None

        return None

    def _parse_product_line(self, line: str, company: Optional[str] = None) -> Optional[Dict]:
        line = line.strip()
        if not line:
            return None

        if any(kw in line.upper() for kw in ["现货", "实单", "开价", "OFFER", "出，", "发，"]):
            return None

        part_match = re.match(r"^([A-Z0-9\-:]+)", line)
        if not part_match:
            return None

        part_number = part_match.group(1)
        if not (re.search(r"[A-Z]", part_number) and re.search(r"\d", part_number)):
            return None

        product = build_empty_product()
        product["企业名称"] = company
        product["料号型号"] = part_number
        product["产品名称"] = part_number
        # 🦞 先搜精选 - 品牌名映射（中文->英文）
        brand = self._identify_brand(part_number)
        product["品牌"] = map_brand_name(brand, part_number)
        product["厂家"] = map_brand_name(brand, part_number)
        product["交货天数"] = 1

        dc_match = re.search(r"(\d{2})\s*\+", line)
        if dc_match:
            product["批次（DC）"] = dc_match.group(1) + "+"

        price_match = re.search(r"(\d+\.?\d*)\s*(?:USD|U)\b", line, re.IGNORECASE)
        if price_match:
            try:
                product["单价"] = float(price_match.group(1))
            except Exception:
                pass

        for keyword, meaning in self.REMARK_KEYWORDS.items():
            if keyword in line:
                product["货物情况（多选）"] = meaning
                break

        if product["单价"] is None:
            return None

        return product

    def _identify_brand(self, part_number: str) -> str:
        for brand, patterns in self.BRAND_PATTERNS.items():
            for pattern in patterns:
                if re.match(pattern, part_number, re.IGNORECASE):
                    return brand
        return "未知品牌"


class UniversalOfferParser:
    """通用 OFFER 解析器"""

    BRAND_PATTERNS = {
        # 英特尔 (Intel) - 优先匹配，避免与 Micron 冲突
        "Intel": [
            r"^29F\d{2,}", r"^PC29", r"^SSD",
        ],
        # 南亚科技 (Nanya)
        "南亚科技 (Nanya)": [
            r"^NT5A", r"^NT5B", r"^NT5C", r"^NT5D", r"^NT5L", r"^NT5U",
            r"^NT6A", r"^NT6B", r"^NT6C", r"^NT6U",
        ],
        # SK Hynix
        "SK Hynix": [
            r"^H5C", r"^H5AN", r"^H5HC", r"^H5T4", r"^H5TQ",
            r"^H9C", r"^H9HC", r"^H9TQ",
        ],
        # 镁光 (Micron) - 29F 已移至 Intel
        "Micron": [
            r"^MT4\b", r"^MT40", r"^MT41", r"^MT41K", r"^MT41J", r"^MT42", r"^MT46",
            r"^MT5\b", r"^MT53", r"^MT53E", r"^MT58",
            r"^MT6\b", r"^MT62", r"^MT62F",
            r"^MT29F",
            r"^25Q",
        ],
        # 三星 (Samsung)
        "Samsung": [
            r"^M32",
            r"^K4\b", r"^K4A", r"^K4B", r"^K4U", r"^K4F", r"^K4G", r"^K4L",
            r"^K8\b", r"^K8E", r"^K8G",
            r"^KLM", r"^KMF", r"^KLUD",
            r"^HN8T", r"^H26T", r"^THG",
            r"^K4AAG", r"^K3K", r"^K3U", r"^KMZ",
        ],
        # 金士顿 (Kingston)
        "Kingston": [
            r"^D\d", r"^D1", r"^D2", r"^D3", r"^D4",
            r"^KVR",
        ],
        # 兆易创新 (GigaDevice)
        "GigaDevice": [
            r"^GD\b", r"^GD25", r"^GD25L", r"^GD25Q", r"^GD25W",
        ],
        # 华邦 (Winbond)
        "Winbond": [
            r"^W\b", r"^W25", r"^W25Q", r"^W25N", r"^W94", r"^W98",
        ],
        # 晶存 (Puya)
        "晶存 (Puya)": [
            r"^P25\b", r"^P25D", r"^P25Q",
        ],
    }

    def __init__(self):
        self.format_detected = None

    def detect_format(self, text: str) -> str:
        lines = text.strip().split("\n")

        # 检测 HK OFFER 格式 (2026-01-08 HK OFFER)
        has_hk_offer_header = any(re.match(r'\d{4}-\d{2}-\d{2}\s+HK\s+OFFER', line, re.IGNORECASE) for line in lines)
        has_brand_category = any(re.match(r'^(SAMSUNG|SK HYNIX|SK Hynix|NANYA|Nanya|MICRON|CXMT)\s+(DDR|LPDDR|EMMC)', line, re.IGNORECASE) for line in lines)
        has_spec_format = any(re.match(r'^\d+\*\d+\s+[A-Z0-9\-]+\s+\$', line) for line in lines)
        
        has_hkstock_title = any(line.strip() == "出，香港现货！" for line in lines)
        has_detailed = any(re.search(r"[A-Z0-9\-]+\s+[A-Z0-9]+\s+\d+GB", line) for line in lines)
        has_pipe = any("|" in line for line in lines)
        has_usd = any("usd" in line.lower() for line in lines)
        has_part_prefix = any(line.lower().startswith("part") for line in lines)
        has_price_prefix = any("price" in line.lower() for line in lines)
        has_comma_separated = any(line.count(",") >= 3 for line in lines)

        if has_hk_offer_header or (has_brand_category and has_spec_format):
            return "hk_offer_format"
        if has_hkstock_title:
            return "hkstock_format"
        if has_detailed:
            return "detailed_format"
        if has_pipe:
            return "table_pipe"
        if has_part_prefix or has_price_prefix:
            return "email_format"
        if has_comma_separated and not has_usd:
            return "csv_format"
        if has_usd:
            return "hongda_format"
        return "short_format"

    def parse(self, text: str) -> List[Dict]:
        self.format_detected = self.detect_format(text)

        if self.format_detected == "hk_offer_format":
            return HKOfferParser().parse(text)
        if self.format_detected == "detailed_format":
            return DetailedLineParser().parse(text)
        if self.format_detected == "hkstock_format":
            return HKStockParser().parse(text)
        if self.format_detected == "hongda_format":
            return OfferParser().parse_offer_text(text)
        if self.format_detected == "table_pipe":
            return self._parse_table_pipe(text)
        if self.format_detected == "email_format":
            return self._parse_email(text)
        if self.format_detected == "csv_format":
            return self._parse_csv(text)
        if self.format_detected == "short_format":
            return self._parse_short(text)
        return OfferParser().parse_offer_text(text)

    def _parse_table_pipe(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        lines = text.strip().split("\n")
        headers = None

        for line in lines:
            line = line.strip()
            if not line:
                continue

            fields = [f.strip() for f in line.split("|")]
            if headers is None:
                headers = fields
                continue

            product = self._fields_to_product(headers, fields)
            if product and product.get("料号型号"):
                products.append(product)

        return products

    def _parse_email(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        blocks = re.split(r"\n\s*\n", text)

        patterns = {
            "料号型号": [
                r"(?:Part(?: Number)?|MPN|Model)[:\s]+([A-Z0-9\-:]+)",
                r"^([A-Z]{2,}\d+[A-Z0-9\-]+)",
            ],
            "品牌": [r"(?:Brand|Mfr|Manufacturer)[:\s]+([A-Za-z\s]+)"],
            "单价": [r"(?:Price|Unit Price|USD)[:\s]+\$?([\d.]+)"],
            "数量": [r"(?:Qty|Quantity|Stock)[:\s]+(\d+)"],
            "批次（DC）": [r"(?:DC|Date Code|Batch)[:\s]+(\d+\+?)"],
        }

        for block in blocks:
            block = block.strip()
            if not block:
                continue

            product = build_empty_product()
            product["企业名称"] = "鸿达半导体"

            for field, pattern_list in patterns.items():
                for pattern in pattern_list:
                    match = re.search(pattern, block, re.IGNORECASE)
                    if not match:
                        continue
                    value = match.group(1).strip()
                    if field == "单价":
                        try:
                            product[field] = float(value)
                        except Exception:
                            pass
                    elif field == "数量":
                        try:
                            product[field] = int(value)
                        except Exception:
                            pass
                    else:
                        product[field] = value
                    break

            if product["料号型号"] and not product["品牌"]:
                product["品牌"] = self._identify_brand(product["料号型号"])
                product["厂家"] = product["品牌"]

            if product["料号型号"]:
                product["产品名称"] = product["料号型号"]
                products.append(product)

        return products

    def _parse_csv(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        lines = text.strip().split("\n")
        headers = None

        for line in lines:
            line = line.strip()
            if not line:
                continue

            fields = [f.strip() for f in line.split(",")]
            if headers is None:
                headers = fields
                continue

            product = self._fields_to_product(headers, fields)
            if product and product.get("料号型号"):
                products.append(product)

        return products

    def _parse_short(self, text: str) -> List[Dict]:
        products: List[Dict] = []
        lines = text.strip().split("\n")

        for line in lines:
            line = line.strip()
            if not line:
                continue

            part_match = re.match(r"^([A-Z]{2,}\d+[A-Z0-9\-:]+)", line)
            if not part_match:
                continue

            part_number = part_match.group(1)
            product = build_empty_product()
            product["企业名称"] = "鸿达半导体"
            product["料号型号"] = part_number
            product["产品名称"] = part_number
            # 🦞 先搜精选 - 品牌名映射（中文->英文）
            brand = self._identify_brand(part_number)
            product["品牌"] = map_brand_name(brand, part_number)
            product["厂家"] = map_brand_name(brand, part_number)

            price_match = re.search(r"\$?([\d.]+)", line)
            if price_match:
                try:
                    product["单价"] = float(price_match.group(1))
                except Exception:
                    pass

            dc_match = re.search(r"(\d{2}\+)", line)
            if dc_match:
                product["批次（DC）"] = dc_match.group(1)

            products.append(product)

        return products

    def _fields_to_product(self, headers: List[str], fields: List[str]) -> Dict:
        product = build_empty_product()
        product["企业名称"] = "鸿达半导体"

        field_map = {
            "料号": "料号型号",
            "料号型号": "料号型号",
            "part": "料号型号",
            "part number": "料号型号",
            "mpn": "料号型号",
            "型号": "料号型号",
            "品牌": "品牌",
            "brand": "品牌",
            "mfr": "品牌",
            "manufacturer": "品牌",
            "单价": "单价",
            "价格": "单价",
            "price": "单价",
            "unit price": "单价",
            "usd": "单价",
            "数量": "数量",
            "qty": "数量",
            "quantity": "数量",
            "stock": "数量",
            "dc": "批次（DC）",
            "批次": "批次（DC）",
            "date code": "批次（DC）",
            "batch": "批次（DC）",
            "货所在地": "货所在地",
            "地点": "货所在地",
            "location": "货所在地",
            "交期": "交货天数",
            "交货期": "交货天数",
            "lead time": "交货天数",
        }

        for i, header in enumerate(headers):
            if i >= len(fields):
                break

            mapped = field_map.get(header.lower().strip(), header.lower().strip())
            if mapped not in product:
                continue

            value = fields[i].strip()
            if not value or value == "-":
                continue

            if mapped == "单价":
                try:
                    value = float(value.replace("$", "").replace(",", ""))
                except Exception:
                    pass
            elif mapped == "数量":
                try:
                    value = int(value.replace(",", ""))
                except Exception:
                    pass

            product[mapped] = value

        if product["料号型号"] and not product["品牌"]:
            # 🦞 先搜精选 - 品牌名映射（中文->英文）
            brand = self._identify_brand(product["料号型号"])
            product["品牌"] = map_brand_name(brand, part_number)
            product["厂家"] = map_brand_name(brand, part_number)

        if product["料号型号"]:
            product["产品名称"] = product["料号型号"]

        return product

    def _identify_brand(self, part_number: str) -> str:
        for brand, patterns in self.BRAND_PATTERNS.items():
            for pattern in patterns:
                if re.match(pattern, part_number, re.IGNORECASE):
                    return brand
        return "未知品牌"


# ============================================================================
# FastAPI backend
# ============================================================================


class LoginRequest(BaseModel):
    mobile: str
    smsCode: str = "2222"


class LoginResponse(BaseModel):
    success: bool
    message: str
    token: Optional[str] = None
    username: Optional[str] = None


class ParseRequest(BaseModel):
    text: str


class ParseResponse(BaseModel):
    success: bool
    count: int
    products: List[Dict]
    message: str


class PublishRequest(BaseModel):
    text: str
    mobile: Optional[str] = None  # 🦞 先搜精选 - 动态手机号
    sms_code: Optional[str] = "2222"  # 🦞 先搜精选 - 动态验证码


class PublishResponse(BaseModel):
    success: bool
    message: str
    excel_path: Optional[str] = None
    count: int


user_sessions: Dict[str, Dict] = {}


def _extract_token_from_auth_header(authorization: Optional[str]) -> Optional[str]:
    if not authorization:
        return None
    auth = authorization.strip()
    if not auth:
        return None
    if auth.lower().startswith("bearer "):
        token = auth[7:].strip()
        return token or None
    return auth


def _find_session_by_token(token: str) -> Optional[Dict]:
    for session in user_sessions.values():
        if session.get("token") == token:
            return session
    return None


app = FastAPI(
    title="自动发布机器人 API",
    description="鸿达半导体自动发布系统（单文件版）",
    version="3.0-single",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.post("/api/login", response_model=LoginResponse)
async def api_login(request: LoginRequest):
    try:
        login_mgr = LoginManager()
        success = login_mgr.login(request.mobile, request.smsCode)

        if not success:
            return LoginResponse(success=False, message="登录失败，请检查手机号和验证码")

        token = login_mgr.get_token()
        user_info = login_mgr.user_info or {}
        user_sessions[request.mobile] = {
            "token": token,
            "session": login_mgr.get_session(),
            "user_info": user_info,
        }

        return LoginResponse(
            success=True,
            message="登录成功",
            token=token,
            username=user_info.get("username"),
        )
    except Exception as e:
        return LoginResponse(success=False, message=f"登录失败:{str(e)}")


@app.post("/api/parse", response_model=ParseResponse)
async def api_parse_offer(request: ParseRequest):
    try:
        parser = UniversalOfferParser()
        products = parser.parse(request.text)

        if not products:
            return ParseResponse(success=False, count=0, products=[], message="未解析到任何产品")

        product_list = []
        for p in products:
            product_list.append(
                {
                    "料号型号": p.get("料号型号", ""),
                    "品牌": p.get("品牌", ""),
                    "单价": p.get("单价", 0),
                    "货币单位": p.get("货币单位（usd/cny）", "usd"),
                    "数量": p.get("数量"),
                    "批次": p.get("批次（DC）"),
                    "交货天数": p.get("交货天数", 7),
                    "货所在地": p.get("货所在地", "HK"),
                    "企业名称": p.get("企业名称", "鸿达半导体"),
                    "货物情况": p.get("货物情况（多选）", "全新"),
                    "产品分类": p.get("产品分类", "存储芯片"),
                }
            )

        return ParseResponse(
            success=True,
            count=len(products),
            products=product_list,
            message=f"成功解析 {len(products)} 个产品",
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"解析失败:{str(e)}") from e


@app.post("/api/publish", response_model=PublishResponse)
async def api_publish(request: PublishRequest, authorization: Optional[str] = Header(None)):
    """🦞 先搜精选 - 支持动态手机号发布"""
    try:
        # 优先使用请求中的手机号，否则使用 token 验证
        mobile = request.mobile or None
        
        if mobile:
            # 🦞 先搜精选 - 使用动态手机号登录
            login_mgr = LoginManager()
            login_success = login_mgr.login(mobile=mobile, sms_code=request.sms_code or "2222")
            
            if not login_success:
                raise HTTPException(status_code=401, detail=f"登录失败（手机号：{mobile}）")
            
            session = login_mgr.get_session()
            headers = login_mgr.get_headers()
        else:
            # 使用旧的 token 验证方式
            token = _extract_token_from_auth_header(authorization)
            if not token:
                raise HTTPException(status_code=401, detail="未授权访问")

            user_session = _find_session_by_token(token)
            if not user_session:
                raise HTTPException(status_code=401, detail="Token 无效或已过期")
            
            session = user_session["session"]
            headers = user_session["session"].headers

        parser = UniversalOfferParser()
        products = parser.parse(request.text)
        if not products:
            raise HTTPException(status_code=400, detail="未解析到任何产品")

        # 🦞 先搜精选 - 品牌名映射
        for p in products:
            if p.get("品牌"):
                p["品牌"] = map_brand_name(p["品牌"])
            if p.get("厂家"):
                p["厂家"] = map_brand_name(p["厂家"])

        generator = ExcelGenerator()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"供需信息_{timestamp}.xlsx"
        filepath = generator.create_excel(products, filename)
        if not filepath:
            raise HTTPException(status_code=500, detail="Excel 生成失败")

        uploader = Uploader(session=session, headers=headers)
        if not mobile:
            uploader.headers["Authorization"] = f"Bearer {token}"

        upload_success = uploader.upload_excel(filepath)
        if not upload_success:
            detail = uploader.last_error or "上传失败"
            raise HTTPException(status_code=500, detail=f"上传失败:{detail}")

        return PublishResponse(
            success=True,
            message=f"成功发布 {len(products)} 个产品",
            excel_path=filepath,
            count=len(products),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"发布失败:{str(e)}") from e


@app.get("/api/health")
async def health_check():
    return {
        "status": "ok",
        "service": "自动发布机器人 API",
        "version": "3.0-single",
        "base_url": BASE_URL,
    }


@app.get("/")
async def root():
    return {
        "message": "欢迎使用自动发布机器人 API（单文件版）",
        "docs": "/docs",
        "health": "/api/health",
    }


# ============================================================================
# Pipelines / CLI
# ============================================================================


def run_main_pipeline() -> bool:
    print("=" * 60)
    print("🤖 自动发布机器人 v1.0（单文件版）")
    print("=" * 60)
    print(f"⏰ 启动时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📊 运行模式：{RUN_MODE}")
    print()

    if RUN_MODE == "REAL":
        print("【步骤 1/4】登录系统")
        print("-" * 60)

        login_mgr = LoginManager()
        # 🦞 先搜精选 v4.1 - 强制检查手机号
        if not mobile:
            print("❌ 错误：必须提供手机号，不支持默认值")
            return False
        if not sms_code:
            print("❌ 错误：请提供短信验证码")
            return False
        if not BASE_URL or not LOGIN_API:
            print("❌ 错误：请配置 BASE_URL 和 LOGIN_API")
            return False
        if not login_mgr.login():
            print("❌ 登录失败，程序终止")
            return False
        print()
    else:
        print("【步骤 1/4】跳过登录（模拟模式）")
        print("-" * 60)
        print("ℹ️  提示：修改 RUN_MODE='REAL' 启用真实上传")
        print()

    print("【步骤 2/4】解析产品信息")
    print("-" * 60)
    parser = ProductParser()
    products = parser.parse_directory()

    if not products:
        print("❌ 没有解析到任何产品，程序终止")
        return False

    print("\n【验证必填字段】")
    required_fields = ["企业名称", "产品名称", "料号型号"]
    valid_products = []

    for i, product in enumerate(products, 1):
        missing = [f for f in required_fields if not product.get(f)]
        if missing:
            print(f"⚠️  产品{i}: 缺少字段 {missing}，已跳过")
        else:
            valid_products.append(product)
            print(f"✅ 产品{i}: {product.get('产品名称')} - 验证通过")

    if not valid_products:
        print("❌ 没有有效产品，程序终止")
        return False

    products = valid_products
    print()

    print("【步骤 3/4】生成 Excel 表格")
    print("-" * 60)
    generator = ExcelGenerator()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"供需信息_{timestamp}.xlsx"
    filepath = generator.create_excel(products, filename)

    if not filepath:
        print("❌ Excel 生成失败，程序终止")
        return False

    print()

    if RUN_MODE == "REAL":
        print("【步骤 4/4】上传发布")
        print("-" * 60)

        uploader = Uploader(session=login_mgr.get_session(), headers=login_mgr.get_headers())
        if not uploader.upload_excel(filepath):
            print("⚠️  上传失败，但 Excel 文件已保存在本地")
            print(f"   文件路径：{filepath}")
            return False

        print()
        print("=" * 60)
        print("✅ 发布完成！")
        print("=" * 60)
    else:
        print("【步骤 4/4】跳过上传（模拟模式）")
        print("-" * 60)
        print("✅ Excel 文件已生成，可手动上传")
        print()
        print("=" * 60)
        print("✅ 模拟运行完成！")
        print("=" * 60)

    print(f"📊 成功处理产品数：{len(products)}")
    print(f"📄 Excel 文件：{filepath}")
    print()
    return True


def run_offer_pipeline(offer_file: Optional[str] = None) -> bool:
    print("=" * 60)
    print("🤖 自动发布机器人 - OFFER 解析版（单文件）")
    print("=" * 60)
    print(f"⏰ 时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    print("【步骤 1/2】解析 OFFER 文本")
    print("-" * 60)

    offer_parser = OfferParser()
    if not offer_file:
        offer_file = str(BASE_DIR / "products" / "offer_20260313.txt")

    if not os.path.exists(offer_file):
        print(f"❌ 文件不存在：{offer_file}")
        return False

    products = offer_parser.parse_offer_file(offer_file)
    if not products:
        print("❌ 没有解析到任何产品")
        return False

    print(f"\n✅ 成功解析 {len(products)} 个产品")
    print()

    print("【步骤 2/2】生成 Excel 表格")
    print("-" * 60)

    generator = ExcelGenerator()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"鸿达_OFFER_{timestamp}.xlsx"
    filepath = generator.create_excel(products, filename)

    if not filepath:
        print("❌ Excel 生成失败")
        return False

    print()
    print("=" * 60)
    print("✅ OFFER 处理完成！")
    print("=" * 60)
    print(f"📊 产品数量：{len(products)}")
    print(f"📄 Excel 文件：{filepath}")
    return True


def run_api_server(host: str = "0.0.0.0", port: int = 8123):
    print("=" * 60)
    print("🤖 自动发布机器人 API 服务（单文件版）")
    print("=" * 60)
    print(f"📡 服务地址:http://{host}:{port}")
    print(f"📚 API 文档:http://{host}:{port}/docs")
    print()
    print("🔧 接口配置:")
    print(f"   BASE_URL: {BASE_URL}")
    print(f"   LOGIN_API: {LOGIN_API}")
    print(f"   UPLOAD_API: {UPLOAD_API}")
    print()
    print("🚀 启动服务...")
    print()

    uvicorn.run(app, host=host, port=port, log_level="info")


def build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="自动发布机器人单文件入口")
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("main", help="运行主流程（读取 products/*.txt）")

    offer = sub.add_parser("offer", help="运行 OFFER 解析流程")
    offer.add_argument("--file", dest="offer_file", default=None, help="OFFER 文本文件路径")

    api = sub.add_parser("api", help="启动 FastAPI 服务")
    api.add_argument("--host", default="0.0.0.0")
    api.add_argument("--port", type=int, default=8123)

    return parser


def main_cli() -> int:
    parser = build_cli()
    args = parser.parse_args()

    if args.command in (None, "main"):
        return 0 if run_main_pipeline() else 1
    if args.command == "offer":
        return 0 if run_offer_pipeline(args.offer_file) else 1
    if args.command == "api":
        run_api_server(args.host, args.port)
        return 0
    parser.print_help()
    return 1


if __name__ == "__main__":
    sys.exit(main_cli())
