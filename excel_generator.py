#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
自动发布机器人 - Excel 表格生成模块
版本：v1.0
日期：2026-03-13
"""

import openpyxl
from openpyxl import Workbook
from typing import List, Dict
import os
import sys
from datetime import datetime
sys.path.append('/Users/mac/Desktop/自动发布机器人')
from config import OUTPUT_DIR


class ExcelGenerator:
    """Excel 表格生成器"""
    
    # Excel 列顺序（标准格式）
    COLUMNS = [
        '企业名称', '货主', '性别', '手机号', '微信', '产品名称', '品牌', '料号型号',
        '单价', '货币单位（usd/cny）', '数量', '单位', '货所在地', '批次（DC）',
        '货物情况（多选）', '产品分类', '应用', '起订量', '有效期', '交货天数',
        '厂家', '重量', '尺寸 - 长', '尺寸 - 宽', '尺寸 - 高', '替代品牌', '替代型号'
    ]
    
    def __init__(self):
        self.wb = None
        self.ws = None
    
    def create_excel(self, products: List[Dict], filename: str = None) -> str:
        """
        创建 Excel 文件
        
        Args:
            products: 产品数据列表
            filename: 输出文件名（可选）
            
        Returns:
            生成的文件路径
        """
        if not products:
            print("❌ 没有产品数据")
            return None
        
        # 创建工作簿
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "供需信息"
        
        # 写入表头
        self._write_headers()
        
        # 写入数据
        self._write_data(products)
        
        # 调整列宽
        self._auto_adjust_columns()
        
        # 保存文件
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"供需信息_{timestamp}.xlsx"
        
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
        
        filepath = os.path.join(OUTPUT_DIR, filename)
        self.wb.save(filepath)
        
        print(f"✅ Excel 文件已生成：{filepath}")
        print(f"   共 {len(products)} 条产品数据")
        
        return filepath
    
    def _write_headers(self):
        """写入表头"""
        for col, header in enumerate(self.COLUMNS, 1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = header
            
            # 设置样式
            cell.font = openpyxl.styles.Font(bold=True, size=11)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            cell.fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True, size=11)
    
    def _write_data(self, products: List[Dict]):
        """写入数据"""
        for row_idx, product in enumerate(products, 2):
            for col_idx, field in enumerate(self.COLUMNS, 1):
                value = product.get(field)
                
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                # 设置对齐
                cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    
    def _auto_adjust_columns(self):
        """自动调整列宽"""
        column_widths = {
            '企业名称': 20,
            '货主': 10,
            '性别': 6,
            '手机号': 13,
            '微信': 15,
            '产品名称': 25,
            '品牌': 15,
            '料号型号': 20,
            '单价': 10,
            '货币单位（usd/cny）': 12,
            '数量': 8,
            '单位': 6,
            '货所在地': 15,
            '批次（DC）': 12,
            '货物情况（多选）': 15,
            '产品分类': 15,
            '应用': 15,
            '起订量': 8,
            '有效期': 12,
            '交货天数': 8,
            '厂家': 15,
            '重量': 8,
            '尺寸 - 长': 8,
            '尺寸 - 宽': 8,
            '尺寸 - 高': 8,
            '替代品牌': 15,
            '替代型号': 15,
        }
        
        for col_idx, header in enumerate(self.COLUMNS, 1):
            width = column_widths.get(header, 12)
            self.ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    def append_to_existing(self, filepath: str, products: List[Dict]) -> bool:
        """
        追加到现有 Excel 文件
        
        Args:
            filepath: 现有文件路径
            products: 产品数据列表
            
        Returns:
            是否成功
        """
        if not os.path.exists(filepath):
            print(f"❌ 文件不存在：{filepath}")
            return False
        
        self.wb = openpyxl.load_workbook(filepath)
        
        if "供需信息" not in self.wb.sheetnames:
            print("❌ 文件中没有'供需信息'工作表")
            return False
        
        self.ws = self.wb["供需信息"]
        
        # 找到最后一行
        start_row = self.ws.max_row + 1
        
        # 写入数据
        for row_idx, product in enumerate(products, start_row):
            for col_idx, field in enumerate(self.COLUMNS, 1):
                value = product.get(field)
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.value = value
        
        self.wb.save(filepath)
        print(f"✅ 已追加 {len(products)} 条数据到：{filepath}")
        return True


if __name__ == '__main__':
    # 测试生成
    from parser import ProductParser
    
    parser = ProductParser()
    products = parser.parse_directory()
    
    if products:
        generator = ExcelGenerator()
        filepath = generator.create_excel(products)
        print(f"\n生成的文件：{filepath}")
